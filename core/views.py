from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages

from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login

from django.views.decorators.http import require_GET
from django.utils import timezone
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from .models import TimeSlot

from .models import Field, Booking
from .forms import ProfileForm

from datetime import datetime
from decimal import Decimal

from django.template.loader import render_to_string
import pdfkit  
import openpyxl
from openpyxl.utils import get_column_letter
from django.core.mail import send_mail
from django.conf import settings


# ============================================================
# PUBLIC PAGES
# ============================================================

def home(request):
    return render(request, 'home.html')


def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Account created successfully!")
            return redirect('home')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = UserCreationForm()

    return render(request, 'register.html', {'form': form})


def field_list(request):
    fields = Field.objects.all()
    return render(request, 'field_list.html', {'fields': fields})


# ============================================================
# BOOKING SYSTEM
# ============================================================


@login_required
def book_field(request, field_id):
    field = get_object_or_404(Field, id=field_id)
    slots = TimeSlot.objects.filter(field=field).order_by('start_time')

    if request.method == 'POST':
        slot_id = request.POST.get('slot_id')
        slot = get_object_or_404(TimeSlot, id=slot_id)

        date = request.POST.get('date')

        # conflict check
        conflict = Booking.objects.filter(
            field=field,
            date=date,
            start_time=slot.start_time,
            end_time=slot.end_time,
            status='approved'
        ).exists()

        if conflict:
            messages.error(request, "Slot already booked.")
            return redirect('book_field', field_id=field.id)

        # compute amount
        duration_hours = (datetime.combine(date=datetime.today(), time=slot.end_time) -
                          datetime.combine(date=datetime.today(), time=slot.start_time)).seconds / 3600

        amount = Decimal(duration_hours) * Decimal(field.price_per_hour)

        Booking.objects.create(
            user=request.user,
            field=field,
            date=date,
            start_time=slot.start_time,
            end_time=slot.end_time,
            status='pending',
            amount=amount,
            payment_status='unpaid'
        )

        messages.success(request, "Booking request submitted!")
        return redirect('my_bookings')

    return render(request, 'book_field.html', {'field': field, 'slots': slots})


@login_required
def my_bookings(request):
    bookings = Booking.objects.filter(user=request.user).order_by('-date', '-start_time')
    return render(request, 'my_bookings.html', {'bookings': bookings})


# ============================================================
# RECEIPTS
# ============================================================

@login_required
def booking_receipt(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id, user=request.user)
    return render(request, 'booking_receipt.html', {'booking': booking})


@staff_member_required
def admin_receipt(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    return render(request, 'booking_receipt.html', {'booking': booking, 'admin_view': True})


@login_required
def booking_receipt_pdf(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id, user=request.user)

    html = render_to_string('booking_receipt.html', {'booking': booking})
    pdf = pdfkit.from_string(html, False)

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=receipt_{booking.id}.pdf'
    return response


# ============================================================
# ADMIN MANAGEMENT
# ============================================================

@staff_member_required
def admin_dashboard(request):
    bookings = Booking.objects.all().order_by('-date', '-start_time')
    return render(request, 'admin_dashboard.html', {'bookings': bookings})


@staff_member_required
def update_booking_status(request, booking_id, status):
    booking = get_object_or_404(Booking, id=booking_id)

    if request.method == 'POST':
        booking.status = status
        booking.save()
        messages.success(request, f"Booking updated to {status.title()}.")

    return redirect('admin_dashboard')


@staff_member_required
def update_payment_status(request, booking_id, action):
    booking = get_object_or_404(Booking, id=booking_id)

    if request.method == 'POST':
        if action == "paid":
            booking.payment_status = "paid"
            booking.payment_date = timezone.now()
        elif action == "unpaid":
            booking.payment_status = "unpaid"
            booking.payment_date = None
        elif action == "refunded":
            booking.payment_status = "refunded"

        booking.save()
        messages.success(request, "Payment status updated.")

    return redirect('admin_dashboard')


# ============================================================
# ANALYTICS DASHBOARD
# ============================================================

@staff_member_required
def analytics_dashboard(request):
    total_revenue = Booking.objects.filter(payment_status='paid').aggregate(Sum('amount'))['amount__sum'] or 0
    total_bookings = Booking.objects.count()
    approved_bookings = Booking.objects.filter(status='approved').count()

    monthly = (
        Booking.objects.filter(payment_status='paid')
        .annotate(month=TruncMonth('date'))
        .values('month')
        .annotate(total=Sum('amount'))
        .order_by('month')
    )

    labels = [m['month'].strftime("%b %Y") for m in monthly]
    data = [float(m['total']) for m in monthly]

    return render(request, 'analytics_dashboard.html', {
        'total_revenue': total_revenue,
        'total_bookings': total_bookings,
        'approved_bookings': approved_bookings,
        'labels': labels,
        'data': data,
    })


# ============================================================
# CALENDAR – SINGLE FIELD
# ============================================================

@login_required
def availability_calendar(request, field_id):
    field = get_object_or_404(Field, id=field_id)
    return render(request, 'availability_calendar.html', {'field': field})


@require_GET
def availability_api(request, field_id):
    field = get_object_or_404(Field, id=field_id)

    qs = Booking.objects.filter(field=field)
    qs = qs.filter(status__in=['approved', 'pending']) if request.user.is_staff else qs.filter(status='approved')

    events = [{
        "id": b.id,
        "title": f"{b.field.name}",
        "start": f"{b.date}T{b.start_time}",
        "end": f"{b.date}T{b.end_time}",
        "color": "#28a745" if b.status == "approved" else "#ffc107",
    } for b in qs]

    return JsonResponse(events, safe=False)


# ============================================================
# CALENDAR – ALL FIELDS COMBINED
# ============================================================

@login_required
def all_fields_calendar(request):
    fields = Field.objects.all()
    return render(request, 'all_fields_calendar.html', {'fields': fields})


@require_GET
def all_fields_api(request):
    fields = Field.objects.all()
    events = []

    colors = ["#1abc9c", "#3498db", "#9b59b6", "#f39c12", "#e74c3c", "#2ecc71", "#34495e"]

    for field in fields:
        qs = Booking.objects.filter(field=field)
        qs = qs.filter(status__in=['approved', 'pending']) if request.user.is_staff else qs.filter(status='approved')

        field_color = colors[(field.id - 1) % len(colors)]

        for b in qs:
            events.append({
                "id": b.id,
                "title": f"{field.name}",
                "start": f"{b.date}T{b.start_time}",
                "end": f"{b.date}T{b.end_time}",
                "color": field_color,
            })

    return JsonResponse(events, safe=False)


# ============================================================
# PROFILE
# ============================================================

@login_required
def profile_view(request):
    if request.method == 'POST':
        form = ProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Profile updated!")
            return redirect('profile')
    else:
        form = ProfileForm(instance=request.user)

    return render(request, 'profile.html', {'form': form})


def field_detail(request, field_id):
    field = get_object_or_404(Field, id=field_id)
    return render(request, 'field_detail.html', {'field': field})

@staff_member_required
def export_bookings_excel(request):
    bookings = Booking.objects.all().order_by('-date', '-start_time')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings"

    # Header row
    headers = [
        "User", "Field", "Date", "Start Time", "End Time", 
        "Amount (Rs)", "Payment Status", "Booking Status", "Created At"
    ]

    ws.append(headers)

    # Data rows
    for b in bookings:
        ws.append([
            b.user.username,
            b.field.name,
            b.date.strftime("%Y-%m-%d"),
            b.start_time.strftime("%H:%M"),
            b.end_time.strftime("%H:%M"),
            float(b.amount),
            b.payment_status,
            b.status,
            b.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    # Auto-column width
    for i, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bookings.xlsx"'

    wb.save(response)
    return response

def send_booking_email(booking, event_type):
    """
    event_type: 'created', 'approved', 'rejected', 'payment'
    """
    user_email = booking.user.email
    if not user_email:
        return  # no email set, just skip

    subject = ""
    message = ""

    if event_type == 'created':
        subject = "Futsal Booking Request Received"
        message = (
            f"Hi {booking.user.username},\n\n"
            f"We have received your booking request for {booking.field.name}.\n"
            f"Date: {booking.date}\nTime: {booking.start_time} - {booking.end_time}\n"
            f"Amount: Rs. {booking.amount}\n\n"
            f"Status: Pending approval.\n\n"
            "Thank you for using our Futsal Management System."
        )
    elif event_type == 'approved':
        subject = "Futsal Booking Approved ✅"
        message = (
            f"Hi {booking.user.username},\n\n"
            f"Your booking for {booking.field.name} has been APPROVED.\n"
            f"Date: {booking.date}\nTime: {booking.start_time} - {booking.end_time}\n"
            f"Amount: Rs. {booking.amount}\n\n"
            "You can view your booking and receipt in your account.\n\n"
            "Thank you!"
        )
    elif event_type == 'rejected':
        subject = "Futsal Booking Rejected ❌"
        message = (
            f"Hi {booking.user.username},\n\n"
            f"Unfortunately, your booking for {booking.field.name} on {booking.date} "
            f"({booking.start_time} - {booking.end_time}) was REJECTED.\n\n"
            "You may try another time slot.\n\n"
            "Thank you."
        )
    elif event_type == 'payment':
        subject = "Payment Status Updated"
        message = (
            f"Hi {booking.user.username},\n\n"
            f"Payment status for your booking ({booking.field.name}, {booking.date} "
            f"{booking.start_time}-{booking.end_time}) is now: {booking.payment_status.upper()}.\n\n"
            "Thank you."
        )

    if subject and message:
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [user_email],
            fail_silently=True,  # avoid crashing if email fails
        )
